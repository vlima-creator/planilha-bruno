import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from utils.metricas import calcular_metricas, calcular_qualidade_arquivo
from utils.analises import analisar_skus, analisar_motivos, analisar_frete

def aplicar_estilo_cabecalho(sheet, columns_count):
    """Aplica estilo profissional ao cabeçalho da planilha"""
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    alignment = Alignment(horizontal='center', vertical='center')
    
    for col in range(1, columns_count + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

def ajustar_largura_colunas(sheet):
    """Ajusta a largura das colunas com base no conteúdo"""
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = min(adjusted_width, 60)

def formatar_valores_excel(sheet, columns_monetarias, columns_percentuais):
    """Aplica formatos numéricos do Excel para valores monetários e percentuais"""
    for row in range(2, sheet.max_row + 1):
        for col in columns_monetarias:
            if col <= sheet.max_column:
                cell = sheet.cell(row=row, column=col)
                cell.number_format = 'R$ #,##0.00'
        for col in columns_percentuais:
            if col <= sheet.max_column:
                cell = sheet.cell(row=row, column=col)
                cell.number_format = '0.00%'

def exportar_xlsx(data):
    """Exporta os resultados para um arquivo XLSX"""
    vendas = data['vendas']
    matriz = data['matriz'] if data['matriz'] is not None else pd.DataFrame()
    full = data['full'] if data['full'] is not None else pd.DataFrame()
    max_date = data['max_date']
    plataforma = data.get('plataforma', 'ML')
    
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    
    # 1. ABA RESUMO EXECUTIVO
    metricas_total = calcular_metricas(vendas, matriz, full, max_date, 180)
    
    resumo_data = [
        ['MÉTRICA DE DESEMPENHO', 'VALOR ATUAL'],
        ['Plataforma', plataforma],
        ['Total de Pedidos Processados', metricas_total['vendas']],
        ['Faturamento Bruto de Produtos', metricas_total['faturamento_produtos']],
        ['Faturamento Total (c/ Fretes)', metricas_total['faturamento_total']],
        ['---', '---'],
        ['Total de Devoluções', metricas_total['devolucoes_vendas']],
        ['Taxa de Devolução Global', metricas_total['taxa_devolucao']],
        ['Faturamento Comprometido (Devoluções)', metricas_total['faturamento_devolucoes']],
        ['---', '---'],
        ['Perda Total Estimada', abs(metricas_total['perda_total'])],
        ['Perda Parcial (Custos Operacionais)', abs(metricas_total['perda_parcial'])],
        ['---', '---'],
        ['Análise de Saúde das Devoluções', 'Quantidade'],
        ['Devoluções Saudáveis', metricas_total['saudaveis']],
        ['Devoluções Críticas', metricas_total['criticas']],
        ['Devoluções Neutras', metricas_total['neutras']],
    ]
    
    df_resumo = pd.DataFrame(resumo_data[1:], columns=resumo_data[0])
    df_resumo.to_excel(writer, sheet_name='Resumo Executivo', index=False)
    
    ws_resumo = writer.sheets['Resumo Executivo']
    aplicar_estilo_cabecalho(ws_resumo, 2)
    # Formatação para valores monetários e percentuais no resumo vertical
    for row_idx, row_val in enumerate(resumo_data, 1):
        label = row_val[0]
        if 'Faturamento' in label or 'Perda' in label:
            ws_resumo.cell(row=row_idx, column=2).number_format = 'R$ #,##0.00'
        if 'Taxa' in label:
            ws_resumo.cell(row=row_idx, column=2).number_format = '0.00%'
    ajustar_largura_colunas(ws_resumo)

    # 2. ABA RANKING DE SKUS (TOP 50)
    df_skus, _ = analisar_skus(vendas, matriz, full, max_date, 180, limit=50)
    if not df_skus.empty:
        df_skus.to_excel(writer, sheet_name='Ranking de SKUs', index=False)
        ws_skus = writer.sheets['Ranking de SKUs']
        aplicar_estilo_cabecalho(ws_skus, len(df_skus.columns))
        # Colunas: SKU, Vendas, Dev, Impacto, Taxa
        # No ranking de SKUs, Impacto é a 4ª coluna e Taxa é a 5ª
        formatar_valores_excel(ws_skus, [4], [5])
        ajustar_largura_colunas(ws_skus)

    # 3. ABA MOTIVOS DE DEVOLUÇÃO
    df_motivos = analisar_motivos(vendas, matriz, full, max_date, 180)
    if not df_motivos.empty:
        df_motivos.to_excel(writer, sheet_name='Motivos de Devolução', index=False)
        ws_motivos = writer.sheets['Motivos de Devolução']
        aplicar_estilo_cabecalho(ws_motivos, len(df_motivos.columns))
        formatar_valores_excel(ws_motivos, [], [3])
        ajustar_largura_colunas(ws_motivos)

    # 4. ABA ANÁLISE DE LOGÍSTICA
    df_frete = analisar_frete(vendas, matriz, full, max_date, 180)
    if not df_frete.empty:
        df_frete.to_excel(writer, sheet_name='Análise de Logística', index=False)
        ws_frete = writer.sheets['Análise de Logística']
        aplicar_estilo_cabecalho(ws_frete, len(df_frete.columns))
        # Colunas: Forma de Entrega, Vendas, Devoluções, Taxa (%), Impacto (R$)
        formatar_valores_excel(ws_frete, [5], [4])
        ajustar_largura_colunas(ws_frete)

    # 5. ABA QUALIDADE DOS DADOS
    qualidade = calcular_qualidade_arquivo(data)
    qualidade_rows = [
        ['INDICADOR DE QUALIDADE', 'PERCENTUAL DE FALHA'],
        ['Vendas: SKUs não identificados', qualidade['vendas']['sem_sku_pct'] / 100],
        ['Vendas: Datas ausentes', qualidade['vendas']['sem_data_pct'] / 100],
        ['Vendas: N.º de venda ausente', qualidade['vendas']['sem_numero_venda_pct'] / 100],
    ]
    if plataforma == 'ML':
        qualidade_rows.extend([
            ['Devoluções Matriz: Sem motivo informado', qualidade['matriz']['sem_motivo_pct'] / 100],
            ['Devoluções Matriz: Sem estado do produto', qualidade['matriz']['sem_estado_pct'] / 100],
            ['Devoluções Full: Sem motivo informado', qualidade['full']['sem_motivo_pct'] / 100],
            ['Devoluções Full: Sem estado do produto', qualidade['full']['sem_estado_pct'] / 100],
        ])
    
    pd.DataFrame(qualidade_rows[1:], columns=qualidade_rows[0]).to_excel(writer, sheet_name='Qualidade dos Dados', index=False)
    ws_qual = writer.sheets['Qualidade dos Dados']
    aplicar_estilo_cabecalho(ws_qual, 2)
    formatar_valores_excel(ws_qual, [], [2])
    ajustar_largura_colunas(ws_qual)

    # 6. ABA DADOS BRUTOS (VENDAS)
    vendas_export = vendas.head(5000)
    vendas_export.to_excel(writer, sheet_name='Base de Vendas', index=False)
    ws_vendas = writer.sheets['Base de Vendas']
    aplicar_estilo_cabecalho(ws_vendas, len(vendas_export.columns))
    ajustar_largura_colunas(ws_vendas)

    # 7. ABA DEVOLUÇÕES (CONSOLIDADO)
    devolucoes_raw = pd.concat([matriz, full], ignore_index=True)
    if not devolucoes_raw.empty:
        devolucoes_raw.to_excel(writer, sheet_name='Base de Devoluções', index=False)
        ws_dev = writer.sheets['Base de Devoluções']
        aplicar_estilo_cabecalho(ws_dev, len(devolucoes_raw.columns))
        ajustar_largura_colunas(ws_dev)

    writer.close()
    output.seek(0)
    return output
